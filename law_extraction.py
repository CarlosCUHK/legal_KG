from bs4 import BeautifulSoup
import re
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, Color
import backoff  
import tiktoken
import asyncio
import numpy as np 
import json
import logging
import openai
from tqdm import tqdm

openai.api_key = "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"
api_key = "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"
class OutOfQuotaException(Exception):
    "Raised when the key exceeded the current quota"
    def __init__(self, key, cause=None):
        super().__init__(f"No quota for key: {key}")
        self.key = key
        self.cause = cause

    def __str__(self):
        if self.cause:
            return f"{super().__str__()}. Caused by {self.cause}"
        else:
            return super().__str__()

class AccessTerminatedException(Exception):
    "Raised when the key has been terminated"
    def __init__(self, key, cause=None):
        super().__init__(f"Access terminated key: {key}")
        self.key = key
        self.cause = cause

    def __str__(self):
        if self.cause:
            return f"{super().__str__()}. Caused by {self.cause}"
        else:
            return super().__str__()

class DocumentExtractor:    
    def __init__(self, filename) -> None:
        self.content_dict = {}
        with open(filename, "r", encoding="utf-8") as file:
            self.content = file.read()
        self.soup = BeautifulSoup(self.content, "html.parser")

    def remove_unnecessary_content_within_parentheses(self, text):
        pattern = r'\([^)]*年[^)]*\)'  # Regular expression pattern to match parentheses containing "年"
        modified_text = re.sub(pattern, '', text)
        pattern = r"\[.*?\]"
        modified_text = re.sub(pattern, "", modified_text)
        modified_text = re.sub(r'\(\s*\)', '', modified_text)
        modified_text = modified_text.replace("— ", "")
        return re.sub(r'\(\)', '', modified_text)       

    def remove_additional_spaces(self, text):
        pattern = r'\s{3,}'  
        modified_text = re.sub(pattern, ' ', text)
        return modified_text

    def remove_starting_parenthesis(self, text):
        pattern = r'^\s*\)'  # Regular expression pattern to match spaces followed by a closing parenthesis at the beginning of the string
        modified_text = re.sub(pattern, '', text)
        return modified_text

    def extract_title(self):
        #element = self.soup.find("h1", class_="cap-title-text")
        element = self.soup.find("span",id="LegDetails")
        self.content_dict["title"] = element.get_text()

    def extract_main_definition(self):
        pattern = r'^P\d+\w*'
        # divs = self.soup.find_all("div", class_="hklm_part", attrs={"name": re.compile(pattern)})
        # if divs:
        #     hklm_def_tags = self.soup.find_all('div', class_='hklm_def')
        #     definition_dict = {}
        #     for tag in hklm_def_tags:
        #         parent = tag.find_parent()
        #         if 'hklm_section' in parent.get('class'):
        #             definition_name = tag.find('div', class_="hklm_term") 
        #             en_term_div = tag.find('div', class_='hklm_term', lang='en')
        #             if en_term_div:
        #                 en_term_div.extract()
        #             definition_name.extract()
        #             content = tag.text.replace('\n', '')
        #             content = self.remove_starting_parenthesis(content)
        #             content = self.remove_additional_spaces(content)
        #             if definition_name:
        #                 definition_dict[definition_name.text.replace('\n', '')] = self.remove_unnecessary_content_within_parentheses(content)
        #     self.content_dict["main_definition"] = definition_dict 
        # else:
        sections = self.soup.find_all("div", class_="hklm_section")
        parent = None
        for section in sections:
            if "釋義" in section.text:
                hklm_def_tags = section.find_all('div', class_='hklm_def')
                definition_dict = {}
                for tag in hklm_def_tags:
                    definition_name = tag.find('div', class_="hklm_term") 
                    en_term_div = tag.find('div', class_='hklm_term', lang='en')
                    if en_term_div:
                        en_term_div.extract()
                    definition_name.extract()
                    content = tag.text.replace('\n', '')
                    content = self.remove_starting_parenthesis(content)
                    content = self.remove_additional_spaces(content)
                    if definition_name:
                        definition_dict[definition_name.text.replace('\n', '')] = self.remove_unnecessary_content_within_parentheses(content)
                self.content_dict["main_definition"] = definition_dict 
                parent = section.find_parent()
                parent = parent.find_parent()
                break
        # next_sibling = parent.find_next_sibling()
        # while next_sibling:
        #     self.content_dict["main_definition"]
        #     next_sibling = parent.find_next_sibling()
        #     point_number = next_sibling.find("div", class_="hklm_num")
        #     if point_number:
        #         point_number = point_number.text.replace('\n', '')
        #         sourceNote = next_sibling.find("div", class_="hklm_sourceNote")
        #         if not next_sibling.find_all("div", class_="hklm_content") and sourceNote and '廢除' in sourceNote:
        #             if next_sibling.find_next_sibling().get('class') != ["hklm_main"]:
        #                 break
        #             else:
        #                 continue
                    
        #         self.content_dict[point_number] = {}
        #         heading = next_sibling.find("div", class_="hklm_heading")
        #         if heading:
        #             self.content_dict[point_number]["heading"] = heading.text.replace('\n', '')
            
        #         self.content_dict[point_number]["content"] = self.extract_subpoints(next_sibling)
        #     next_sibling = next_sibling.find_next_sibling()
        #     if "第II部" in next_sibling.text:
        #         break
    

    def extract_subsection_document(self):
        pattern = r'^P\d+\w*'
        divs = self.soup.find_all("div", class_="hklm_part", attrs={"name": re.compile(pattern)})
        if divs:
            for div in divs:
                subsection_dict = {}
                if div.get('name') != 'P1':
                    chapter = div.find("div", class_="hklm_num")
                    chapter_name = div.find("div", class_="hklm_heading")
                    if chapter_name:
                        subsection_dict['chapter_name'] = chapter_name.text.replace('\n', '')
                    else:
                        subsection_dict['chapter_name'] = "None"
                    parent = div.find_parent()
                    next_sibling = parent.find_next_sibling()
                    while next_sibling:
                        if next_sibling.find("div", class_="hklm_part", attrs={"name": re.compile(pattern)}):
                            break
                        point_number = next_sibling.find("div", class_="hklm_num")
                        if point_number:
                            point_number = point_number.text.replace('\n', '')
                            sourceNote = next_sibling.find("div", class_="hklm_sourceNote")
                            if not next_sibling.find_all("div", class_="hklm_content") and sourceNote and '廢除' in sourceNote:
                                if next_sibling.find_next_sibling().get('class') != ["hklm_main"]:
                                    break
                                else:
                                    continue
                                
                            subsection_dict[point_number] = {}
                            heading = next_sibling.find("div", class_="hklm_heading")
                            if heading:
                                subsection_dict[point_number]["heading"] = heading.text.replace('\n', '')
                        
                            subsection_dict[point_number]["content"] = self.extract_subpoints(next_sibling)
                            next_sibling = next_sibling.find_next_sibling()
                            
                        if next_sibling.get('class') != ["hklm_main"]:
                            break
                    self.content_dict[chapter.text.replace('\n', '')] = subsection_dict
        else:
            divs = self.soup.find_all("div", class_="hklm_crossHeading add-toc")
            for div in divs:
                subsection_dict = {}
                if div.text != '導言':
                    subsection_dict['chapter_name'] = div.text.replace('\n', '')
                    parent = div.find_parent()
                    next_sibling = parent.find_next_sibling()
                    while next_sibling:
                        if next_sibling.find("div", class_="hklm_crossHeading add-toc"):
                            break
                        point_number = next_sibling.find("div", class_="hklm_num")
                        if point_number:
                            point_number = point_number.text.replace('\n', '')
                            sourceNote = next_sibling.find("div", class_="hklm_sourceNote")
                            if not next_sibling.find_all("div", class_="hklm_content") and sourceNote and '廢除' in sourceNote:
                                if next_sibling.find_next_sibling().get('class') != ["hklm_main"]:
                                    break
                                else:
                                    continue
                                
                            subsection_dict[point_number] = {}
                            heading = next_sibling.find("div", class_="hklm_heading")
                            if heading:
                                subsection_dict[point_number]["heading"] = heading.text.replace('\n', '')
                        
                            subsection_dict[point_number]["content"] = self.extract_subpoints(next_sibling)
                            next_sibling = next_sibling.find_next_sibling()
                            
                        if next_sibling.get('class') != ["hklm_main"]:
                            break
                        if not next_sibling.find_all("div", class_="hklm_section"):
                            break
                        
                    self.content_dict[div.text.replace('\n', '')] = subsection_dict
                    
    def extract_subpoints(self, point):
        point_dict = {}
        if not point.find("div", class_="hklm_subsection"): 
            if point.find_all('div', class_='hklm_def'):
                sub_definitions = point.find_all('div', class_='hklm_def')
                sub_definition_dict = {}
                pattern = r'\(\)'
                for sub_definition in sub_definitions:
                    definition_name = sub_definition.find('div', class_="hklm_term")
                    if definition_name:
                        definition = definition_name.text
                        definition_name.extract()
                        definition_en = sub_definition.find('div', class_='hklm_term', lang='en') 
                        if definition_en:
                            definition_en.extract()
                        content = sub_definition.text.replace('\n', '')
                        content = self.remove_additional_spaces(content)
                        content = self.remove_unnecessary_content_within_parentheses(content)
                        sub_definition_dict[definition.replace('\n', '')] = re.sub(pattern, '', content)
                    point_dict["content"] = sub_definition_dict 
            else:
                if point.find("div", class_="hklm_content"): 
                    content = point.text
                    content = self.remove_additional_spaces(content)
                    content = self.remove_unnecessary_content_within_parentheses(content)
                    point_dict["content"] = content.replace('\n', '')
                else:
                    return None
        else:
            subsections = point.find_all("div", class_="hklm_subsection")
            for subsection in subsections:
                content = ""
                order = subsection.find("div", class_="hklm_num no_heading_follows")
                if order:
                    if subsection.find_all("div", class_="hklm_content"):
                        content = subsection.text
                        content = self.remove_additional_spaces(content)
                        content = self.remove_unnecessary_content_within_parentheses(content)  
                        point_dict[order.text.replace('\n', '').replace('“', '')] = content.replace('\n', '')
                    else:
                        sub_definitions = subsection.find_all('div', class_='hklm_def')
                        sub_definition_dict = {}
                        pattern = r'\(\)'
                        for sub_definition in sub_definitions:
                            definition_name = sub_definition.find('div', class_="hklm_term")
                            definition = definition_name.text
                            definition_name.extract()
                            definition_en = sub_definition.find('div', class_='hklm_term', lang='en') 
                            if definition_en:
                                definition_en.extract()
                            content = sub_definition.text.replace('\n', '')
                            content = self.remove_additional_spaces(content)
                            content = self.remove_unnecessary_content_within_parentheses(content)
                            sub_definition_dict[definition.replace('\n', '').replace('“', '')] = re.sub(pattern, '', content)
                        point_dict[order.text.replace('\n', '').replace('“', '')] = sub_definition_dict 
        return point_dict
    
    @backoff.on_exception(backoff.expo, (openai.error.RateLimitError, openai.error.APIError, openai.error.ServiceUnavailableError, openai.error.APIConnectionError), max_tries=5)
    async def translate_with_backoff(self, messages, len_prompt, api_key, temperature=0.05):
        try:
            response = await openai.ChatCompletion.acreate(
                model='gpt-3.5-turbo',
                messages=messages,
                temperature=temperature,
                top_p=1.0,
                max_tokens=4000-len_prompt,
                api_key=api_key,
            )
            gen = response['choices'][0]['message']['content'].strip().replace('\n\n\n', '\n\n')
            
            if gen == "":
                gen = " "
                
            return gen
        except openai.error.RateLimitError as e:
            if "You exceeded your current quota, please check your plan and billing details" in e.user_message:
                raise OutOfQuotaException(api_key)
            elif "Your access was terminated due to violation of our policies" in e.user_message:
                raise AccessTerminatedException(api_key)
            else:
                raise e
            
    def get_messages(self, prompt):
        messages = []
        prompt = prompt.split('\n\n\n')
        for id, smp in enumerate(prompt):
            if id == 0:
                messages.append(
                        {"role": "system", "content": smp}
                    )
            else:
                messages.append(
                    {"role": "user", "content": smp}
                )
        return messages

    def num_tokens_from_string(self, string: str, encoding_name: str) -> int:
        """Returns the number of tokens in a text string."""
        encoding = tiktoken.get_encoding(encoding_name)
        num_tokens = len(encoding.encode(string))
        return num_tokens
    
    def generate_embedding(self, text:str):
        v = openai.Embedding.create(input=[text], model='text-embedding-ada-002')['data'][0]['embedding']
        return np.array(v)
    

    async def paraphrase(self):
        for key, value in tqdm(self.content_dict.items()):
            if key == "title":
                continue
            elif key == "main_definition":
                continue
                # if self.content_dict["main_definition"]:
                #     iterator = iter(self.content_dict["main_definition"])
                #     for i in range(len(self.content_dict["main_definition"])):
                #         definition = next(iterator)
                #         message = "以一段, 對下文進行概括。直接以”當、凡、若、除...“等條件句開始" + "\n\n\n" + self.content_dict["main_definition"][definition]
                #         messages = self.get_messages(message)
                #         len_prompt = self.num_tokens_from_string(message, encoding_name="p50k_base") # wrong encoding
                #         try:
                #             gen = await self.translate_with_backoff(
                #                 messages=messages,
                #                 len_prompt=len_prompt,
                #                 api_key=api_key,
                #                 temperature=0
                #             )
                #             self.content_dict["main_definition"][definition] = gen
                #             print(gen)
                #         except (OutOfQuotaException) as e:
                #             logging.warning(e)
                #             return
                #         except openai.error.OpenAIError as e:
                #             await asyncio.sleep(10)
            else:
                chapter_dict = self.content_dict[key]                
                for local_key, local_value in chapter_dict.items():
                    if local_key != "chapter_name":
                        point_dict = chapter_dict[local_key]
                        if point_dict:
                            for subpoint_key, subpoint_value in point_dict.items():
                                if subpoint_key == "heading":
                                    continue
                                elif subpoint_value:
                                    small_point_dict = point_dict[subpoint_key]
                                    for smallsub_key, smallsub_value in small_point_dict.items():
                                        if smallsub_value and type(smallsub_value) == dict:
                                            continue
                                            # sub_def_dict = smallsub_value
                                            # for sub_def_key, sub_def_value in sub_def_dict.items():
                                            #     message = "以一段, 對下文進行概括。直接以”當、凡、若、除...“等條件句開始" + "\n\n\n" + self.content_dict[key][local_key][subpoint_key][smallsub_key][sub_def_key]
                                            #     messages = self.get_messages(message)
                                            #     len_prompt = self.num_tokens_from_string(message, encoding_name="p50k_base") # wrong encoding
                                            # try:
                                            #     gen = await self.translate_with_backoff(
                                            #         messages=messages,
                                            #         len_prompt=len_prompt,
                                            #         api_key=api_key,
                                            #         temperature=0
                                            #     )
                                            #     self.content_dict[key][local_key][subpoint_key][smallsub_key][sub_def_key] = gen
                                            # except (OutOfQuotaException) as e:
                                            #     logging.warning(e)
                                            #     return
                                            # except openai.error.OpenAIError as e:
                                            #     await asyncio.sleep(10)
                                        elif smallsub_value: 
                                            if smallsub_key == "content":
                                                message = "以一段, 對下文進行概括。直接以”當、凡、若、除...“等條件句開始" + "\n\n\n" + smallsub_value
                                                messages = self.get_messages(message)
                                                len_prompt = self.num_tokens_from_string(message, encoding_name="p50k_base") # wrong encoding
                                                try:
                                                    gen = await self.translate_with_backoff(
                                                        messages=messages,
                                                        len_prompt=len_prompt,
                                                        api_key=api_key,
                                                        temperature=0
                                                    )
                                                    self.content_dict[key][local_key][subpoint_key][smallsub_key] = gen
                                                    print(gen)
                                                except (OutOfQuotaException) as e:
                                                    logging.warning(e)
                                                    return
                                                except openai.error.OpenAIError as e:
                                                    await asyncio.sleep(10)
                                            else:
                                                message = f"對下文進行整體閲讀，並以一段對下文第{smallsub_key}大項進行概括。直接以”當、凡、若、除...“等條件句開始" + "\n\n\n" 
                                                for tmp_key, tmp_value in small_point_dict.items():
                                                    if type(tmp_value) != dict:
                                                        message += tmp_value + "  "
                                                messages = self.get_messages(message)
                                                len_prompt = self.num_tokens_from_string(message, encoding_name="p50k_base") # wrong encoding
                                                try:
                                                    gen = await self.translate_with_backoff(
                                                        messages=messages,
                                                        len_prompt=len_prompt,
                                                        api_key=api_key,
                                                        temperature=0
                                                    )
                                                    self.content_dict[key][local_key][subpoint_key][smallsub_key] = gen
                                                    print(gen)
                                                except (OutOfQuotaException) as e:
                                                    logging.warning(e)
                                                    return
                                                except openai.error.OpenAIError as e:
                                                    await asyncio.sleep(10)      
                                                         
    def convert_content_to_embedding(self):
        title = self.content_dict["title"]
        for key, value in tqdm(self.content_dict.items()):
            if key == "title":
                continue
            elif key == "main_definition":
                continue
            else:
                chapter_dict = self.content_dict[key]
                for local_key, local_value in chapter_dict.items():
                    if local_key != "chapter_name":
                        point_dict = self.content_dict[key][local_key]
                        if point_dict:
                            for subpoint_key, subpoint_value in point_dict.items():
                                if subpoint_key == "heading":
                                    self.content_dict[key][local_key][subpoint_key] = self.generate_embedding(self.content_dict[key][local_key][subpoint_key])           
                                    print(self.content_dict[key][local_key][subpoint_key])
                                elif subpoint_value:
                                    small_point_dict = point_dict[subpoint_key]
                                    for smallsub_key, smallsub_value in small_point_dict.items():
                                        if smallsub_value and type(smallsub_value) == dict:
                                            pass
                                        elif smallsub_value: 
                                            self.content_dict[key][local_key][subpoint_key][smallsub_key] = self.generate_embedding(self.content_dict[key][local_key][subpoint_key][smallsub_key])   
                                      
        
def print_nested_dict(dictionary, indent=''):
    for key, value in dictionary.items():
        if isinstance(value, dict):
            print(f"{indent}{key}:\n")
            print_nested_dict(value, indent + '  ')
        else:
            print(f"{indent}{key}: {value}\n")

def print_nested_dict_to_text(dictionary, indent='', file=None, spacing_multiplier=1):
    for key, value in dictionary.items():
        if isinstance(value, dict):
            file.write(f"{indent}{key}:\n")
            print_nested_dict_to_text(value, indent + '  ', file, spacing_multiplier * 0.8)
            file.write('\n' * int(spacing_multiplier))
        else:
            file.write(f"{indent}{key}: {value}\n")


def save_nested_dict_to_xlsx(dictionary, sheet):
    title = dictionary["title"]
    for key, value in dictionary.items():
        if key == "title":
            continue
        elif key == "main_definition":
            cell = sheet.cell(row=1, column=1)
            cell.value = "Title"
            font = Font(size=14, bold=True) 
            cell.font = font
            cell = sheet.cell(row=1, column=2)
            cell.value = title
            font = Font(size=14, color=Color(rgb="0000FF"), bold=True)
            cell.font = font
            cell = sheet.cell(row=3, column=1).value = "def"
            if dictionary["main_definition"]:
                iterator = iter(dictionary["main_definition"])
                for i in range(len(dictionary["main_definition"])):
                    definition = next(iterator)
                    sheet.cell(row=3+i, column=2).value = definition
                    sheet.cell(row=3+i, column=3).value = dictionary["main_definition"][definition]
        else:
            current_sheet = workbook.create_sheet(key)
            chapter_dict = dictionary[key]
            cell = current_sheet.cell(row=1, column=1)
            cell.value = "Title"
            font = Font(size=14) 
            cell.font = font
            cell = current_sheet.cell(row=1, column=2)
            cell.value = chapter_dict["chapter_name"]
            font = Font(size=14, bold=True) 
            cell.font = font
            row_num = 3
            for local_key, local_value in chapter_dict.items():
                
                if local_key != "chapter_name":
                    point_dict = chapter_dict[local_key]
                    if point_dict:
                        current_sheet.cell(row=row_num, column=1).value = local_key
                        for subpoint_key, subpoint_value in point_dict.items():
                            if subpoint_key == "heading":
                                current_sheet.cell(row=row_num, column=2).value = "title"
                                cell = current_sheet.cell(row=row_num, column=3)
                                cell.value = subpoint_value
                                row_num += 1
                                font = Font(bold=True) 
                                cell.font = font
                            elif subpoint_value:
                                small_point_dict = point_dict[subpoint_key]
                                for smallsub_key, smallsub_value in small_point_dict.items():
                                    if smallsub_value and type(smallsub_value) == dict:
                                        sub_def_dict = smallsub_value
                                        current_sheet.cell(row=row_num, column=2).value = smallsub_key + "sub-def"
                                        for sub_def_key, sub_def_value in sub_def_dict.items():
                                            current_sheet.cell(row=row_num, column=3).value = sub_def_key
                                            current_sheet.cell(row=row_num, column=4).value = sub_def_value
                                            row_num += 1
                                    elif smallsub_value: 
                                        current_sheet.cell(row=row_num, column=2).value = smallsub_key
                                        current_sheet.cell(row=row_num, column=3).value = smallsub_value
                                        row_num += 1
                            else:
                                row_num += 1
                row_num += 1

       
if __name__ == "__main__":
    workbook = Workbook()

    async def main():
        file = r"C:\Users\94399\Documents\internship\law_document_extraction\57章.html"
        extractor = DocumentExtractor(file)
        extractor.extract_title()
        extractor.extract_main_definition()
        extractor.extract_subsection_document()
        #await extractor.paraphrase()
        #extractor.convert_content_to_embedding()        
        file_path = "57章.json"

        # Open the file in write mode and save the nested dictionary as JSON
        with open(file_path, "w") as json_file:
            json.dump(extractor.content_dict, json_file)
        #print_nested_dict(extractor.content_dict)
        sheet = workbook.active
        new_name = 'Info'
        sheet.title = new_name
        save_nested_dict_to_xlsx(extractor.content_dict, sheet)
        # Define the XLSX file path
        xlsx_file = '57章.xlsx'
        with open("57章.json", 'r', encoding='utf-8') as my_file:
            for line in my_file:
                data = json.loads(line)
                print(data)
        # Save the workbook to the XLSX file
        workbook.save(xlsx_file)
        print("The law document has been saved successfully!!!")
    
    asyncio.run(main())
